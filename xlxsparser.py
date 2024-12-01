import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class parse():
    def __init__(self, item):
        self.item = item
    
    def process(self):
        items = self.item
        # Save item to Excel file using openpyxl
        wb = Workbook()
        wb.remove(wb.active)
        # ws = wb.active

        for item_index, item in enumerate(items):
            ws = wb.create_sheet(title=f"Sheet{item_index + 1}")
            
            for col_num, (key, val) in enumerate(item.items(), 1):
                if key == 'overview':
                    ws.cell(row=1, column=col_num, value=key) #header
                    for idx, itm in enumerate(val, 2):
                        ws.cell(row=idx, column=col_num, value=itm)
                    continue
                if key == 'productselection':
                    ws.cell(row=1, column=col_num, value='product selection') #header
                    ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + 1)
                    ws.cell(row=2, column=col_num, value='Catalog Number') # subheader
                    ws.cell(row=2, column=col_num + 1, value='Description') # subheader
                    if not val:
                        continue
                    for idx, (prod_key, prod_val) in enumerate(val.items(), 3):
                        ws.cell(row=idx, column=col_num, value=prod_key)
                        ws.cell(row=idx, column=col_num + 1, value=prod_val)
                    continue
                if key == 'documents':
                    col_num = col_num + 1
                    ws.cell(row=1, column=col_num, value=key) #header
                    ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + 3)
                    row_idx = 3
                    for doc_item in val:
                        for idx, (doc_key, doc_val) in enumerate(doc_item.items()):
                            if doc_key == 'links':
                                ws.cell(row=2, column=col_num + idx, value='language') # sub header
                                ws.cell(row=2, column=col_num + idx + 1, value='links') # sub header
                                for doc_val_key, doc_val_val in doc_val.items():
                                    ws.cell(row=row_idx, column=col_num + idx, value=doc_val_key)
                                    ws.cell(row=row_idx, column=col_num + idx + 1, value=doc_val_val)
                                    row_idx += 1
                                continue
                            ws.cell(row=2, column=col_num + idx, value=doc_key) # sub header
                            ws.cell(row=row_idx, column=col_num + idx, value=doc_val)
                    continue
                if isinstance(val, list):
                    ws.cell(row=1, column=col_num, value=key) #header
                    for idx, itm in enumerate(val, 2):
                        ws.cell(row=idx, column=col_num, value=itm)
                    continue
                ws.cell(row=1, column=col_num, value=key) #header
                ws.cell(row=2, column=col_num, value=val)

            # Set the width of every cell to 30
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

            # Make the first row bold and in capital letters
            bold_font = Font(bold=True)
            center_alignment = Alignment(horizontal='center')

            # Enable text wrapping for all cells
            wrap_alignment = Alignment(wrap_text=True)

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = wrap_alignment

            for cell in ws[1]:
                if cell.value is not None:
                    cell.value = cell.value.upper()  # Convert to uppercase
                cell.font = bold_font  # Make bold
                cell.alignment = center_alignment  # Center text



            # Make the second row (subheader) bold, in capital letters, and centered
            for cell in ws[2]:
                if cell.value not in ['Catalog Number', 'Description', 'resource', 'publicationnumber', 'links', 'language']:
                    continue
                if cell.value is not None:
                    cell.value = cell.value.upper()  # Convert to uppercase
                cell.font = bold_font  # Make bold
                cell.alignment = center_alignment  # Center text


            # Save the workbook
            wb.save('item.xlsx')